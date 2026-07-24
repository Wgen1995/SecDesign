#!/usr/bin/env python3
"""Convert security_tests.json to Excel tracking sheet."""
import json
import sys
import argparse
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def convert(tests_file: str, output_file: str) -> None:
    with open(tests_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = "安全检查点清单"

    headers = ["检查点ID", "检查点描述", "关联威胁", "CWE", "OWASP", "严重度", "状态", "责任人", "计划完成", "实际完成", "验证结果"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    for test in data.get("test_cases", []):
        cwe_raw = test.get("cwe", "")
        if isinstance(cwe_raw, list):
            cwe_str = ", ".join(cwe_raw)
        else:
            cwe_str = cwe_raw
        owasp_raw = test.get("owasp", "")
        if isinstance(owasp_raw, list):
            owasp_str = ", ".join(owasp_raw)
        else:
            owasp_str = owasp_raw
        ws.append([
            test.get("checklist_id", ""),
            test.get("test_name", ""),
            test.get("threat_id", ""),
            cwe_str,
            owasp_str,
            test.get("severity", ""),
            "⬜待修复",
            "", "", "", ""
        ])

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["G"].width = 12

    wb.save(output_file)
    print(f"Excel written to {output_file} ({ws.max_row - 1} rows)")


def main():
    parser = argparse.ArgumentParser(description="Convert security_tests.json to Excel")
    parser.add_argument("--tests", required=True, help="Path to security_tests.json")
    parser.add_argument("--output", required=True, help="Output Excel file path")
    args = parser.parse_args()
    convert(args.tests, args.output)


if __name__ == "__main__":
    main()
